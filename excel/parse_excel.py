import openpyxl

def split_and_fill_merged_cells(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    for ws in wb.worksheets:
        merged_info = []
        for merged_range in list(ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            value = ws.cell(row=min_row, column=min_col).value
            merged_info.append((min_row, min_col, max_row, max_col, value))
            ws.unmerge_cells(str(merged_range))
        for min_row, min_col, max_row, max_col, value in merged_info:
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value
    wb.save(output_path)

# 用法示例
split_and_fill_merged_cells('./before.xlsx', './after.xlsx')